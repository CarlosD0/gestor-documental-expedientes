from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone
import os
from django.conf import settings
from django.http import FileResponse
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.db.models import Q, Count, IntegerField
from django.db.models.functions import Cast
from django.core.paginator import Paginator
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl import load_workbook
from django.db.models import Count



from .models import (
    Libro,
    Autor,
    Configuracion,
    PerfilUsuario,
    ProgramaBeca,
    Municipio,
    Escuela,
    CABB,
    BitacoraMovimiento,
    Localidad
)
from .forms import (
    LibroForm,
    EntregaOperativaForm,
    UsuarioEditarForm,
    UsuarioOperativoForm,
    AutorForm,
    PerfilUpdateForm,
    ImportarLibrosForm,
    ConfiguracionForm,
    MunicipioForm,
    LocalidadForm,
    ImportarLocalidadesForm,
    EscuelaForm,
    CABBForm,
    ImportarCurpForm,
    ImportarEscuelasForm
)
from .utils import render_to_pdf


def es_bibliotecario(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def obtener_config():
    config, created = Configuracion.objects.get_or_create(id=1)
    return config


def registrar_movimiento(request, accion, descripcion='', libro=None):

    BitacoraMovimiento.objects.create(
        usuario=request.user if request.user.is_authenticated else None,
        libro=libro,
        accion=accion,
        descripcion=descripcion
    )



@user_passes_test(es_bibliotecario)
def ajustes_sistema(request):
    config = obtener_config()

    if request.method == 'POST':
        form = ConfiguracionForm(request.POST, instance=config)

        if form.is_valid():
            form.save()
            messages.success(request, '¡Configuración actualizada correctamente!')
            return redirect('home')

    else:
        form = ConfiguracionForm(instance=config)

    return render(request, 'admin/configuracion.html', {'form': form})


@login_required
def home(request):

    config = obtener_config()

    periodo_seleccionado = request.GET.get(
        'periodo',
        request.session.get('periodo_seleccionado', '')
    )

    request.session['periodo_seleccionado'] = periodo_seleccionado

    periodos = Libro.objects.filter(
        activo=True
    ).exclude(
        periodo_entrega__isnull=True
    ).exclude(
        periodo_entrega=''
    ).values_list(
        'periodo_entrega',
        flat=True
    ).distinct().order_by(
        'periodo_entrega'
    )

    expedientes = Libro.objects.filter(
        activo=True
    )

    if periodo_seleccionado:
        expedientes = expedientes.filter(
            periodo_entrega=periodo_seleccionado
        )

    total_expedientes = expedientes.count()

    total_entregados = expedientes.filter(
        estatus='ENTREGADO'
    ).count()

    total_pendientes = expedientes.filter(
        estatus='PENDIENTE'
    ).count()

    programas_resumen = expedientes.values(
        'programa_beca__nombre'
    ).annotate(
        total=Count('id')
    ).order_by(
        'programa_beca__nombre'
    )

    municipios_resumen = expedientes.exclude(
        municipio__isnull=True
    ).exclude(
        municipio=''
    ).values(
        'municipio'
    ).annotate(
        total=Count('id')
    ).order_by(
        '-total'
    )[:10]

    expedientes_recientes = expedientes.order_by('-id')[:5]

    labels_programas = [
        item['programa_beca__nombre'] or 'Sin programa'
        for item in programas_resumen
    ]

    data_programas = [
        item['total']
        for item in programas_resumen
    ]

    labels_municipios = [
        item['municipio']
        for item in municipios_resumen
    ]

    data_municipios = [
        item['total']
        for item in municipios_resumen
    ]

    context = {
        'es_admin': request.user.is_staff,
        'config': config,

        'periodos': periodos,
        'periodo_seleccionado': periodo_seleccionado,

        'total_expedientes': total_expedientes,
        'total_entregados': total_entregados,
        'total_pendientes': total_pendientes,

        'programas_resumen': programas_resumen,
        'expedientes_recientes': expedientes_recientes,

        'grafica_estatus': [
            total_entregados,
            total_pendientes
        ],

        'labels_programas': labels_programas,
        'data_programas': data_programas,

        'labels_municipios': labels_municipios,
        'data_municipios': data_municipios,
    }

    return render(request, 'home.html', context)


@login_required
def perfil(request):

    if request.method == 'POST':

        form = PerfilUpdateForm(
            request.POST,
            instance=request.user
        )

        if form.is_valid():
            form.save()
            messages.success(
                request,
                'Perfil actualizado correctamente'
            )
            return redirect('perfil')

        else:
            messages.error(
                request,
                'Revisa los datos del formulario'
            )

    else:

        form = PerfilUpdateForm(instance=request.user)

    return render(request, 'usuarios/perfil.html', {
        'form': form
    })

@login_required
def generar_respaldo_bd(request):

    import subprocess
    import tempfile
    

    if not request.user.is_staff:
        return HttpResponse(
            'No tienes permisos para generar respaldos.',
            status=403
        )

    db = settings.DATABASES['default']

    nombre_bd = db.get('NAME')
    usuario_bd = db.get('USER')
    password_bd = db.get('PASSWORD')
    host_bd = db.get('HOST') or 'localhost'
    puerto_bd = db.get('PORT') or '3306'

    fecha = timezone.now().strftime('%Y%m%d_%H%M%S')
    nombre_archivo = f'respaldo_sare_{fecha}.sql'

    archivo_temporal = tempfile.NamedTemporaryFile(
        delete=False,
        suffix='.sql'
    )

    archivo_temporal.close()

    comando = [
        r'C:\xampp\mysql\bin\mysqldump.exe',
        f'-h{host_bd}',
        f'-P{puerto_bd}',
        f'-u{usuario_bd}',
        *( [f'-p{password_bd}'] if password_bd else [] ),
        nombre_bd
    ]

    try:

        with open(archivo_temporal.name, 'w', encoding='utf-8') as salida:
            resultado = subprocess.run(
                comando,
                stdout=salida,
                stderr=subprocess.PIPE,
                text=True
            )

        if resultado.returncode != 0:
            return HttpResponse(
                f'Error al generar respaldo: {resultado.stderr}',
                status=500
            )

        return FileResponse(
            open(archivo_temporal.name, 'rb'),
            as_attachment=True,
            filename=nombre_archivo
        )

    except Exception as e:

        return HttpResponse(
            f'Error inesperado al generar respaldo: {str(e)}',
            status=500
        )


@user_passes_test(es_bibliotecario)
def productividad(request):

    config = obtener_config()
    hoy = timezone.localdate()

    periodo_seleccionado = request.session.get(
        'periodo_seleccionado',
        ''
    )

    entregas = Libro.objects.filter(
        estatus='ENTREGADO',
        autor__isnull=False,
        fecha_entrega__isnull=False
    )

    if periodo_seleccionado:
        entregas = entregas.filter(
            periodo_entrega=periodo_seleccionado
        )

    entregas_hoy = entregas.filter(
        fecha_entrega=hoy
    ).count()

    total_entregas = entregas.count()

    meta_diaria = config.meta_diaria_entregas or 1

    porcentaje_meta = round(
        (entregas_hoy / meta_diaria) * 100
    ) if meta_diaria > 0 else 0

    usuarios_productividad = entregas.values(
        'autor__nombre'
    ).annotate(
        total_capturas=Count('id'),
        capturas_hoy=Count(
            'id',
            filter=Q(fecha_entrega=hoy)
        )
    ).order_by(
        '-total_capturas'
    )

    municipios = Libro.objects.filter(
        activo=True
    )

    if periodo_seleccionado:
        municipios = municipios.filter(
            periodo_entrega=periodo_seleccionado
        )

    municipios = municipios.values(
        'municipio'
    ).annotate(
        total=Count('id'),
        entregados=Count(
            'id',
            filter=Q(estatus='ENTREGADO')
        ),
        pendientes=Count(
            'id',
            filter=Q(estatus='PENDIENTE')
        )
    ).order_by('municipio')

    ultimas_capturas = entregas.order_by(
        '-fecha_entrega'
    )[:10]

    labels_usuarios = [
        item['autor__nombre'] for item in usuarios_productividad
    ]

    data_usuarios = [
        item['total_capturas'] for item in usuarios_productividad
    ]

    context = {
        'config': config,
        'periodo_seleccionado': periodo_seleccionado,
        'total_entregas': total_entregas,
        'entregas_hoy': entregas_hoy,
        'meta_diaria': meta_diaria,
        'porcentaje_meta': porcentaje_meta,
        'usuarios_productividad': usuarios_productividad,
        'municipios': municipios,
        'ultimas_capturas': ultimas_capturas,
        'labels_usuarios': labels_usuarios,
        'data_usuarios': data_usuarios,
    }

    return render(request, 'productividad.html', context)

@login_required
def mi_productividad(request):

    config = obtener_config()
    hoy = timezone.localdate()

    nombre_responsable = (
        f"{request.user.first_name} "
        f"{request.user.last_name}"
    ).strip()

    entregas = Libro.objects.filter(
        estatus='ENTREGADO',
        autor__nombre__iexact=nombre_responsable,
        fecha_entrega__isnull=False
    )

    total_entregas = entregas.count()

    entregas_hoy = entregas.filter(
        fecha_entrega=hoy
    ).count()

    entregas_mes = entregas.filter(
        fecha_entrega__year=hoy.year,
        fecha_entrega__month=hoy.month
    ).count()

    meta_diaria = config.meta_diaria_entregas or 1

    porcentaje_meta = round(
        (entregas_hoy / meta_diaria) * 100
    ) if meta_diaria > 0 else 0

    ultimas_capturas = entregas.order_by(
        '-fecha_entrega'
    )[:10]

    context = {
        'config': config,
        'total_entregas': total_entregas,
        'entregas_hoy': entregas_hoy,
        'entregas_mes': entregas_mes,
        'meta_diaria': meta_diaria,
        'porcentaje_meta': porcentaje_meta,
        'ultimas_capturas': ultimas_capturas,
    }

    return render(request, 'mi_productividad.html', context)


@user_passes_test(es_bibliotecario)
def lista_usuarios(request):

    usuarios = User.objects.all().order_by('username')

    return render(request, 'usuarios/lista_usuarios.html', {
        'usuarios': usuarios
    })


@user_passes_test(es_bibliotecario)
def crear_usuario(request):

    if request.method == 'POST':
        form = UsuarioOperativoForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, 'Usuario operativo creado correctamente.')
            return redirect('lista_usuarios')

    else:
        form = UsuarioOperativoForm()

    return render(request, 'usuarios/form_usuario.html', {
        'form': form,
        'titulo': 'Crear Usuario Operativo'
    })


@user_passes_test(es_bibliotecario)
def editar_usuario(request, id):

    usuario = get_object_or_404(User, id=id)

    if request.method == 'POST':

        form = UsuarioEditarForm(
            request.POST,
            instance=usuario
        )

        if form.is_valid():
            form.save()
            messages.success(
                request,
                'Usuario actualizado correctamente.'
            )
            return redirect('lista_usuarios')

    else:

        form = UsuarioEditarForm(instance=usuario)

    return render(request, 'usuarios/form_usuario.html', {
        'form': form,
        'titulo': 'Editar Usuario'
    })


@user_passes_test(es_bibliotecario)
def activar_usuario(request, id):

    usuario = get_object_or_404(User, id=id)

    if usuario.is_superuser:
        messages.error(request, 'No puedes desactivar al super administrador.')
        return redirect('lista_usuarios')

    usuario.is_active = not usuario.is_active
    usuario.save()

    messages.success(request, 'Estado del usuario actualizado correctamente.')
    return redirect('lista_usuarios')


@login_required
def lista_municipios(request):

    municipios = Municipio.objects.annotate(
    clave_numero=Cast('clave', IntegerField())
     ).order_by('clave_numero')

    return render(request, 'municipios/lista_municipios.html', {
        'municipios': municipios
    })


@user_passes_test(es_bibliotecario)
def registrar_municipio(request):

    if request.method == 'POST':

        form = MunicipioForm(request.POST)

        if form.is_valid():
            form.save()

            messages.success(
                request,
                'Municipio registrado correctamente'
            )

            return redirect('lista_municipios')

    else:
        form = MunicipioForm()

    return render(request, 'municipios/form_municipio.html', {
        'form': form
    })


@login_required
def lista_localidades(request):

    busqueda = request.GET.get("buscar")

    localidades = Localidad.objects.select_related(
        'municipio'
    ).all().order_by(
        'municipio__nombre',
        'clave'
    )

    if busqueda:
        localidades = localidades.filter(
            Q(nombre__icontains=busqueda) |
            Q(clave__icontains=busqueda) |
            Q(municipio__nombre__icontains=busqueda) |
            Q(municipio__clave__icontains=busqueda)
        )

    return render(request, 'localidades/lista_localidades.html', {
        'localidades': localidades,
        'busqueda': busqueda
    })

from django.http import JsonResponse


@login_required
def lista_escuelas(request):

    busqueda = request.GET.get('buscar')

    escuelas = Escuela.objects.select_related(
        'municipio',
        'localidad'
    ).order_by('nombre')

    if busqueda:

        escuelas = escuelas.filter(
            Q(cct__icontains=busqueda) |
            Q(nombre__icontains=busqueda)
        )

    return render(request, 'catalogos/lista_escuelas.html', {
        'escuelas': escuelas
    })


@login_required
def localidades_por_municipio(request, municipio_id):

    localidades = Localidad.objects.filter(
        municipio_id=municipio_id
    ).order_by('clave')

    data = []

    for loc in localidades:
        data.append({
            'id': loc.id,
            'nombre': f'{loc.clave} - {loc.nombre}'
        })

    return JsonResponse(data, safe=False)


@user_passes_test(es_bibliotecario)
def registrar_localidad(request):

    if request.method == 'POST':

        form = LocalidadForm(request.POST)

        if form.is_valid():
            form.save()

            messages.success(
                request,
                'Localidad registrada correctamente'
            )

            return redirect('lista_localidades')

    else:
        form = LocalidadForm()

    return render(request, 'localidades/form_localidad.html', {
        'form': form
    })

@login_required
def escuelas_por_localidad(request, localidad_id):

    escuelas = Escuela.objects.filter(
        localidad_id=localidad_id,
        activo=True
    ).order_by('cct')

    data = []

    for escuela in escuelas:
        data.append({
            'id': escuela.id,
            'cct': escuela.cct,
            'nombre': f'{escuela.cct} - {escuela.nombre}'
        })

    return JsonResponse(data, safe=False)


@login_required
def detalle_escuela_json(request, escuela_id):

    escuela = get_object_or_404(
        Escuela,
        id=escuela_id
    )

    data = {
        'cct': escuela.cct,
        'nombre': escuela.nombre,
        'municipio': escuela.municipio.nombre,
        'localidad': escuela.localidad.nombre,
    }

    return JsonResponse(data)

@user_passes_test(es_bibliotecario)
def importar_localidades_excel(request):

    if request.method == 'POST':

        form = ImportarLocalidadesForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            try:

                archivo = request.FILES['archivo_excel']

                wb = openpyxl.load_workbook(archivo)
                ws = wb.active

                importadas = 0
                errores = 0

                for row in ws.iter_rows(min_row=2, values_only=True):

                    try:

                        clave_municipio = str(row[0]).strip()
                        nombre_municipio = str(row[1]).strip()

                        clave_localidad = str(row[2]).strip()
                        nombre_localidad = str(row[3]).strip()

                        municipio, created = Municipio.objects.get_or_create(
                            clave=clave_municipio,
                            defaults={
                                'nombre': nombre_municipio
                            }
                        )

                        existe = Localidad.objects.filter(
                            municipio=municipio,
                            clave=clave_localidad
                        ).exists()

                        if not existe:

                            Localidad.objects.create(
                                municipio=municipio,
                                clave=clave_localidad,
                                nombre=nombre_localidad
                            )

                            importadas += 1

                    except:
                        errores += 1

                messages.success(
                    request,
                    f'Se importaron {importadas} localidades. '
                    f'Errores: {errores}'
                )

                return redirect('lista_localidades')

            except Exception as e:

                messages.error(
                    request,
                    f'Error al importar archivo: {e}'
                )

    else:

        form = ImportarLocalidadesForm()

    return render(
        request,
        'localidades/importar_localidades.html',
        {
            'form': form
        }
    )

@user_passes_test(es_bibliotecario)
def importar_escuelas_excel(request):

    if request.method == 'POST':

        form = ImportarEscuelasForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            archivo = request.FILES['archivo_excel']

            wb = load_workbook(archivo)
            ws = wb.active

            insertados = 0
            duplicados = 0
            errores = 0

            for row in ws.iter_rows(min_row=2, values_only=True):

                try:

                    clave_municipio = str(row[0]).strip() if row[0] else ""
                    clave_localidad = str(row[2]).strip().zfill(4) if row[2] else ""

                    cct = str(row[4]).strip().upper() if row[4] else ""

                    nombre_escuela = str(row[5]).strip().upper() if row[5] else ""

                    nivel = str(row[6]).strip().upper() if row[6] else ""

                    if not clave_municipio or not clave_localidad or not cct:
                        errores += 1
                        continue

                    municipio = Municipio.objects.filter(
                        clave=clave_municipio
                    ).first()

                    localidad = Localidad.objects.filter(
                        municipio=municipio,
                        clave=clave_localidad
                    ).first()

                    if not municipio or not localidad:
                        errores += 1
                        continue

                    if Escuela.objects.filter(
                        cct__iexact=cct
                    ).exists():

                        duplicados += 1
                        continue

                    Escuela.objects.create(
                        municipio=municipio,
                        localidad=localidad,
                        cct=cct,
                        nombre=nombre_escuela,
                        nivel=nivel,
                        activo=True
                    )

                    insertados += 1

                except Exception:
                    errores += 1

            messages.success(
                request,
                f'Escuelas importadas: {insertados} | '
                f'Duplicadas: {duplicados} | '
                f'Errores: {errores}'
            )

            return redirect('lista_escuelas')

    else:

        form = ImportarEscuelasForm()

    return render(request, 'catalogos/importar_escuelas.html', {
        'form': form
    })

    
@login_required
def estadisticas_escuelas(request):

    municipio_id = request.GET.get('municipio')

    municipios = Municipio.objects.annotate(
        clave_numero=Cast('clave', IntegerField())
    ).order_by('clave_numero')

    total_escuelas = Escuela.objects.count()
    total_municipios = Municipio.objects.count()

    totales_municipio = Escuela.objects.values(
        'municipio__id',
        'municipio__nombre'
    ).annotate(
        total_escuelas=Count('id')
    ).order_by(
        'municipio__nombre'
    )

    municipio_mayor = totales_municipio.order_by(
        '-total_escuelas'
    ).first()

    labels_municipios = [
        item['municipio__nombre'] for item in totales_municipio
    ]

    data_municipios = [
        item['total_escuelas'] for item in totales_municipio
    ]

    estadisticas = None
    municipio_seleccionado = None
    total_seleccionado = 0
    labels_niveles = []
    data_niveles = []

    if municipio_id:

        municipio_seleccionado = Municipio.objects.filter(
            id=municipio_id
        ).first()

        estadisticas = Escuela.objects.filter(
            municipio_id=municipio_id
        ).values(
            'nivel'
        ).annotate(
            total=Count('id')
        ).order_by(
            'nivel'
        )

        total_seleccionado = Escuela.objects.filter(
            municipio_id=municipio_id
        ).count()

        labels_niveles = [
            item['nivel'] for item in estadisticas
        ]

        data_niveles = [
            item['total'] for item in estadisticas
        ]

    return render(request, 'catalogos/estadisticas_escuelas.html', {
        'municipios': municipios,
        'totales_municipio': totales_municipio,
        'estadisticas': estadisticas,
        'municipio_seleccionado': municipio_seleccionado,
        'total_seleccionado': total_seleccionado,
        'municipio_id': municipio_id,

        'total_escuelas': total_escuelas,
        'total_municipios': total_municipios,
        'municipio_mayor': municipio_mayor,
        'labels_municipios': labels_municipios,
        'data_municipios': data_municipios,
        'labels_niveles': labels_niveles,
        'data_niveles': data_niveles,
    })


@login_required
def caratulas(request):

    municipios = Municipio.objects.exclude(
      clave__isnull=True
        ).exclude(
      nombre__isnull=True
       ).annotate(
     clave_numero=Cast('clave', IntegerField())
       ).order_by('clave_numero')

    localidades = Localidad.objects.select_related(
        'municipio'
    ).all().order_by(
        'municipio__nombre',
        'nombre'
    )

    periodos = Libro.objects.filter(
        activo=True
    ).exclude(
        periodo_entrega__isnull=True
    ).exclude(
        periodo_entrega=''
    ).values_list(
        'periodo_entrega',
        flat=True
    ).distinct().order_by(
        'periodo_entrega'
    )


    return render(request, 'caratulas/caratulas.html', {
        'municipios': municipios,
        'localidades': localidades,
        'periodos': periodos,
    })


@login_required
def generar_caratula_pdf(request):

    municipio_id = request.GET.get('municipio')
    localidad_id = request.GET.get('localidad')
    cct = request.GET.get('cct')
    periodo = request.GET.get('periodo')
    if not municipio_id and not localidad_id and not cct and not periodo:

      messages.warning(
        request,
        'Selecciona al menos un filtro para generar la carátula.'
      )

      return redirect('caratulas')
    titulo_caratula = request.GET.get(
    'titulo_caratula',
    'INTEGRACIÓN DE KIT DOCUMENTAL ENTREGA DE TARJETAS 3 NIVELES'
     )

    libros = Libro.objects.filter(
        activo=True,
        estatus='ENTREGADO'
    )

    municipio = None
    localidad = None

    if municipio_id:
        municipio = Municipio.objects.get(id=municipio_id)

        libros = libros.filter(
            municipio__icontains=municipio.nombre
        )

    if localidad_id:
        localidad = Localidad.objects.get(id=localidad_id)

        libros = libros.filter(
            localidad__icontains=localidad.nombre
        )

    if cct:
        libros = libros.filter(
            cct_escuela__icontains=cct
        )

    # FILTRO POR PERIODO
    if periodo:
        libros = libros.filter(
            periodo_entrega__icontains=periodo
        )


    if not libros.exists():

      messages.warning(
        request,
        'No existen expedientes entregados con los filtros seleccionados.'
      )

      return redirect('caratulas')

    libros = libros.order_by(
        'apellido_paterno',
        'apellido_materno',
        'nombre'
    )

    libros_procesados = []

    for libro in libros:

        libros_procesados.append({
            'id_tarjeta': libro.id_tarjeta,
            'curp': libro.curp,
            'nombre': libro.nombre,
            'ape_pat': libro.apellido_paterno,
            'ape_mat': libro.apellido_materno,
            'estatus': libro.estatus,
            'nombre_escuela': libro.nombre_escuela,
        })

    context = {
        'titulo_caratula': titulo_caratula,
        'municipio': municipio,
        'localidad': localidad,
        'cct': cct,
        'libros': libros_procesados,
        'total': len(libros_procesados),
        'fecha': timezone.now()
    }

    return render_to_pdf(
        'caratulas/caratula_pdf.html',
        context
    )


@login_required
def lista_libros(request):

    busqueda = request.GET.get("buscar")
    estado = request.GET.get("estado")
    programa_id = request.GET.get("programa")

    periodo_seleccionado = request.session.get(
        'periodo_seleccionado',
        ''
    )

    programas = ProgramaBeca.objects.filter(
        activo=True
    ).order_by('nombre')

    libros_list = Libro.objects.select_related(
       'programa_beca',
        'autor'
    )

    if not request.user.is_staff:
      libros_list = libros_list.filter(
        activo=True
      )

    libros_list = libros_list.order_by('-id')

    if periodo_seleccionado:
        libros_list = libros_list.filter(
            periodo_entrega=periodo_seleccionado
        )

    if busqueda:
        libros_list = libros_list.filter(
            Q(titulo__icontains=busqueda) |
            Q(curp__icontains=busqueda) |
            Q(id_tarjeta__icontains=busqueda) |
            Q(cct_escuela__icontains=busqueda) |
            Q(nombre_escuela__icontains=busqueda) |
            Q(cupo__icontains=busqueda) |
            Q(municipio__icontains=busqueda) |
            Q(localidad__icontains=busqueda) |
            Q(autor__nombre__icontains=busqueda) |
            Q(programa_beca__nombre__icontains=busqueda)
        ).distinct()

    if estado:
        libros_list = libros_list.filter(
            estatus=estado
        )

    if programa_id:
        libros_list = libros_list.filter(
            programa_beca_id=programa_id
        )

    paginator = Paginator(libros_list, 10)
    page_obj = paginator.get_page(
        request.GET.get('page')
    )

    return render(request, 'expedientes/lista_libros.html', {
        'page_obj': page_obj,
        'programas': programas,
        'programa_id': programa_id,
        'estado': estado,
    })

@login_required
def detalle_libro(request, id):

    libro = get_object_or_404(Libro, id=id)

    movimientos = libro.movimientos.select_related(
        'usuario'
    ).all()

    return render(request, 'expedientes/detalle_libro.html', {
        'libro': libro,
        'movimientos': movimientos
    })

@user_passes_test(es_bibliotecario)
def registrar_libro(request):
    if request.method == 'POST':
        form = LibroForm(request.POST, request.FILES)

        if form.is_valid():
            expediente = form.save(commit=False)


            expediente.periodo_entrega = request.POST.get(
                'periodo_entrega',
                expediente.periodo_entrega
            )

            cupo = expediente.cupo.strip() if expediente.cupo else ""
            perfil = None

            if cupo:

              perfil = PerfilUsuario.objects.filter(
                cupo=cupo
              ).select_related('usuario').first()
            if perfil:
                nombre_responsable = (
                  f"{perfil.usuario.first_name} "
                  f"{perfil.usuario.last_name}"
                ).strip()
                responsable, creado = Autor.objects.get_or_create(
                   nombre=nombre_responsable
                )
                expediente.autor = responsable

            if expediente.cupo and expediente.fecha_entrega and expediente.portada:
                expediente.estatus = 'ENTREGADO'
                expediente.disponible = False
                expediente.capturado_por = request.user
                expediente.fecha_captura_entrega = timezone.now()
            else:
                expediente.estatus = 'PENDIENTE'
                expediente.disponible = True

            expediente.save()

            registrar_movimiento(
               request,
               'CREAR',
               'Registró un nuevo expediente',
               expediente
            )

            messages.success(request, 'Expediente registrado correctamente')
            return redirect('lista_libros')

        else:
            messages.error(request, 'Revisa los campos del formulario.')

    else:
        form = LibroForm()

        municipios = Municipio.objects.exclude(
            clave__isnull=True
        ).exclude(
            nombre__isnull=True
        ).annotate(
            clave_numero=Cast('clave', IntegerField())
        ).order_by('clave_numero')

    return render(request, 'expedientes/form_libro.html', {
        'form': form,
        'titulo': 'Registrar Expediente',
        'municipios': municipios,
    })


@login_required
def editar_libro(request, id):

    libro = get_object_or_404(Libro, id=id)

    if request.user.is_staff:
        FormularioEditar = LibroForm
        plantilla = 'expedientes/form_libro.html'
        titulo = 'Editar Expediente'
    else:
        FormularioEditar = EntregaOperativaForm
        plantilla = 'expedientes/form_entrega_operativa.html'
        titulo = 'Capturar Entrega de Tarjeta'

    municipios = Municipio.objects.exclude(
        clave__isnull=True
    ).exclude(
        nombre__isnull=True
    ).annotate(
        clave_numero=Cast('clave', IntegerField())
    ).order_by('clave_numero')

    if request.method == 'POST':

        form = FormularioEditar(
            request.POST,
            request.FILES,
            instance=libro
        )

        if form.is_valid():

            expediente = form.save(commit=False)

            cupo = expediente.cupo.strip() if expediente.cupo else ""
            perfil = None

            if cupo:
                perfil = PerfilUsuario.objects.filter(
                    cupo=cupo
                ).select_related('usuario').first()

            if perfil:
                nombre_responsable = (
                    f"{perfil.usuario.first_name} "
                    f"{perfil.usuario.last_name}"
                ).strip()

                responsable, creado = Autor.objects.get_or_create(
                    nombre=nombre_responsable
                )

                expediente.autor = responsable

            if expediente.cupo and expediente.fecha_entrega and expediente.portada:

                expediente.estatus = 'ENTREGADO'
                expediente.disponible = False
                expediente.capturado_por = request.user
                expediente.fecha_captura_entrega = timezone.now()

            else:

                expediente.estatus = 'PENDIENTE'
                expediente.disponible = True
                expediente.capturado_por = None
                expediente.fecha_captura_entrega = None

            expediente.save()

            registrar_movimiento(
               request,
               'EDITAR',
               'Editó un expediente',
                libro
            )

            messages.success(
                request,
                'Expediente actualizado correctamente'
            )

            return redirect('lista_libros')

        else:

            messages.error(
                request,
                'Revisa los campos del formulario.'
            )

    else:

        form = FormularioEditar(instance=libro)

    return render(request, plantilla, {
        'form': form,
        'libro': libro,
        'titulo': titulo,
        'municipios': municipios,
    })


@user_passes_test(es_bibliotecario)
def activar_libro(request, id):
    libro = get_object_or_404(Libro, id=id)

    libro.activo = not libro.activo
    libro.save()

    if libro.activo:
        messages.success(request, 'Expediente habilitado correctamente')
    else:
        messages.warning(request, 'Expediente desactivado')

    return redirect('lista_libros')


@user_passes_test(es_bibliotecario)
def eliminar_libro(request, id):
    expediente = get_object_or_404(Libro, id=id)

    if expediente.portada:
        expediente.portada.delete(save=False)

    expediente.delete()

    messages.success(request, 'Expediente eliminado correctamente.')
    return redirect('lista_libros')


@login_required
def marcar_entregado(request, id):
    libro = get_object_or_404(Libro, id=id)

    libro.estatus = 'ENTREGADO'

    if not libro.fecha_entrega:
        libro.fecha_entrega = date.today()

    libro.disponible = False
    libro.capturado_por = request.user
    libro.fecha_captura_entrega = timezone.now()
    libro.save()

    registrar_movimiento(
       request,
       'ENTREGAR',
       'Marcó expediente como entregado',
       libro
    )

    messages.success(request, f'Tarjeta entregada a: {libro.titulo}')
    return redirect('lista_libros')


@user_passes_test(es_bibliotecario)
def reporte_libros_pdf(request):
    config = obtener_config()
    libros = Libro.objects.filter(activo=True).order_by('titulo')

    data = {
        'libros': libros,
        'fecha': timezone.now(),
        'institucion': config.nombre_institucion
    }

    return render_to_pdf('expedientes/reporte_pdf.html', data)



@user_passes_test(es_bibliotecario)
def lista_autores(request):
    autores = Autor.objects.all().order_by('nombre')

    return render(request, 'autores/lista_autores.html', {
        'autores': autores
    })


@user_passes_test(es_bibliotecario)
def registrar_autor(request):
    if request.method == 'POST':
        form = AutorForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(request, 'Registrado')
            return redirect('lista_autores')

    else:
        form = AutorForm()

    return render(request, 'autores/form_autor.html', {
        'form': form
    })


@user_passes_test(es_bibliotecario)
def exportar_libros_excel(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expedientes"

    encabezados = [
        "ID",
        "CURP",
        "ID TARJETA",
        "NOMBRE",
        "APELLIDO PATERNO",
        "APELLIDO MATERNO",
        "PROGRAMA",
        "CCT ESCUELA",
        "NOMBRE ESCUELA",
        "MUNICIPIO",
        "LOCALIDAD",
        "CUPO",
        "PERIODO",
        "FECHA ENTREGA",
        "ESTATUS",
        "RESPONSABLE",
        "CAPTURADO POR",
        "FECHA CAPTURA",
        "OBSERVACIONES"
    ]

    ws.append(encabezados)

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="0d6efd",
            end_color="0d6efd",
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    for l in Libro.objects.all():

        ws.append([
            l.id,
            l.curp,
            l.id_tarjeta,
            l.nombre,
            l.apellido_paterno,
            l.apellido_materno,
            l.programa_beca.nombre if l.programa_beca else "",
            l.cct_escuela,
            l.nombre_escuela,
            l.municipio,
            l.localidad,
            l.cupo,
            l.periodo_entrega,
            str(l.fecha_entrega) if l.fecha_entrega else "",
            l.estatus,
            l.autor.nombre if l.autor else "",
            l.capturado_por.username if l.capturado_por else "",
            l.fecha_captura_entrega.strftime("%d/%m/%Y %H:%M")
            if l.fecha_captura_entrega else "",
            l.descripcion
        ])

    columnas = {
        'A': 10,
        'B': 25,
        'C': 22,
        'D': 25,
        'E': 25,
        'F': 25,
        'G': 25,
        'H': 22,
        'I': 45,
        'J': 25,
        'K': 25,
        'L': 15,
        'M': 25,
        'N': 18,
        'O': 18,
        'P': 30,
        'Q': 25,
        'R': 22,
        'S': 45,
    }

    for col, width in columnas.items():
        ws.column_dimensions[col].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename="Reporte_Expedientes.xlsx"'
    )

    wb.save(response)

    return response


@user_passes_test(es_bibliotecario)
def importar_libros_excel(request):

    if request.method == 'POST':
        form = ImportarLibrosForm(request.POST, request.FILES)

        if form.is_valid():
            try:
                archivo = request.FILES['archivo_excel']

                wb = openpyxl.load_workbook(archivo)
                ws = wb.active

                

                importados = 0
                duplicados = 0
                errores = 0

                # FORMATO NUEVO:
                # CURP | ID TARJETA | NOMBRE | APELLIDO PATERNO | APELLIDO MATERNO |
                # PROGRAMA | CUPO | CCT | ESCUELA | MUNICIPIO | LOCALIDAD |
                # PERIODO | FECHA ENTREGA | ESTATUS | OBSERVACIONES

                for row in ws.iter_rows(min_row=2, values_only=True):

                    try:
                        curp = str(row[0]).strip() if row[0] else ""
                        id_tarjeta = str(row[1]).strip() if row[1] else ""
                        numero_tarjeta_bancaria = str(row[2]).strip() if row[2] else ""

                        nombre = str(row[3]).strip() if row[3] else ""
                        apellido_paterno = str(row[4]).strip() if row[4] else ""
                        apellido_materno = str(row[5]).strip() if row[5] else ""

                        programa_excel = str(row[6]).strip().upper() if row[6] else ""
                        cupo = str(row[7]).strip() if row[7] else ""
                        cct = str(row[8]).strip().upper() if row[8] else ""
                        escuela = str(row[9]).strip().upper() if row[9] else ""
                        municipio = str(row[10]).strip().upper() if row[10] else ""
                        localidad = str(row[11]).strip().upper() if row[11] else ""

                        # AUTOCOMPLETAR DATOS DESDE CCT

                        if cct:

                            escuela_obj = Escuela.objects.filter(
                               cct__iexact=cct
                            ).select_related(
                                'municipio',
                                'localidad'
                            ).first()

                            if escuela_obj:

                                if not escuela:
                                   escuela = escuela_obj.nombre

                                if not municipio:
                                    municipio = escuela_obj.municipio.nombre

                                if not localidad:
                                    localidad = escuela_obj.localidad.nombre
                        periodo_valor = row[12]

                        if periodo_valor:
                             if hasattr(periodo_valor, 'strftime'):
                                periodo = periodo_valor.strftime("%b-%Y").upper()
                             else:
                              periodo = str(periodo_valor).strip().upper()
                        else:
                                periodo = ""
                        fecha_entrega = row[13] if row[13] else None

                        estatus_excel = (
                            str(row[14]).strip().upper()
                            if row[14]
                            else "PENDIENTE"
                        )

                        observaciones = (
                            str(row[15]).strip()
                            if row[15]
                            else ""
                        )

                        if not nombre or not apellido_paterno:
                            errores += 1
                            continue

                        if curp and Libro.objects.filter(curp=curp).exists():
                            duplicados += 1
                            continue

                        if id_tarjeta and Libro.objects.filter(id_tarjeta=id_tarjeta).exists():
                            duplicados += 1
                            continue

                        programa_beca = ProgramaBeca.objects.filter(
                           clave__iexact=programa_excel,
                           activo=True
                        ).first()

                        if not programa_beca:
                          errores += 1
                          continue

                        estatus = "ENTREGADO" if "ENTREGADO" in estatus_excel else "PENDIENTE"

                        responsable = None

                        perfil = None

                        if cupo:

                          perfil = PerfilUsuario.objects.filter(
                             cupo=cupo
                          ).select_related('usuario').first()

                        if perfil:

                            nombre_responsable = (
                               f"{perfil.usuario.first_name} "
                               f"{perfil.usuario.last_name}"
                            ).strip()
                            responsable, creado = Autor.objects.get_or_create(
                              nombre=nombre_responsable
                            )

                        Libro.objects.create(
                           
                            

                            nombre=nombre,
                            apellido_paterno=apellido_paterno,
                            apellido_materno=apellido_materno,
                            curp=curp if curp else None,
                            id_tarjeta=id_tarjeta if id_tarjeta else None,
                            numero_tarjeta_bancaria=numero_tarjeta_bancaria if numero_tarjeta_bancaria else None,
                            programa_beca=programa_beca, 
                            cupo=cupo if cupo else None,
                            cct_escuela=cct,
                            nombre_escuela=escuela,
                            municipio=municipio,
                            localidad=localidad,
                            periodo_entrega=periodo,
                            fecha_entrega=fecha_entrega,
                            estatus=estatus,
                            descripcion=observaciones,
                            autor=responsable,
                            tipo_captura='HISTORICA',
                            activo=True,
                            disponible=False if estatus == "ENTREGADO" else True,
                            capturado_por=request.user if estatus == "ENTREGADO" else None,
                            fecha_captura_entrega=timezone.now() if estatus == "ENTREGADO" else None
                        )

                        importados += 1

                    except Exception as e:
                        errores += 1
                        print("ERROR EN FILA:", row)
                        print("DETALLE:", e)
                registrar_movimiento(
                   request,
                    'IMPORTAR',
                    f'Importó {importados} expedientes desde Excel'
                )        

                messages.success(
                    request,
                    f'Importación finalizada: {importados} importados, {duplicados} duplicados y {errores} con error.'
                )

                return redirect('lista_libros')

            except Exception as e:
                messages.error(
                    request,
                    f'Error al importar archivo: {e}'
                )

    else:
        form = ImportarLibrosForm()

    return render(request, 'expedientes/importar_libros.html', {
        'form': form
    })


@user_passes_test(es_bibliotecario)
def descargar_plantilla_excel(request):

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla Expedientes"

    encabezados = [
        "CURP",
        "ID TARJETA",
        "NUMERO_TARJETA_BANCARIA",
        "NOMBRE",
        "APELLIDO PATERNO",
        "APELLIDO MATERNO",
        "PROGRAMA",
        "CUPO",
        "CCT",
        "ESCUELA",
        "MUNICIPIO",
        "LOCALIDAD",
        "PERIODO",
        "FECHA ENTREGA",
        "ESTATUS",
        "OBSERVACIONES"
    ]

    ws.append(encabezados)

    ws.append([
        "AAAA000000HAAAAA00",
        "TARJETA001",
        "5261960723422232",
        "NOMBRE",
        "APELLIDO PATERNO",
        "APELLIDO MATERNO",
        "RITA_SEC",
        "308458",
        "30DPR0000A",
        "NOMBRE DE LA ESCUELA",
        "PAPANTLA",
        "LOCALIDAD",
        "MAYO-AGOSTO 2026",
        "",
        "PENDIENTE",
        "OBSERVACIÓN OPCIONAL"
    ])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="198754",
            end_color="198754",
            fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center")

    columnas = {
        'A': 25,
        'B': 20,
        'C': 30,
        'D': 25,
        'E': 25,
        'F': 18,
        'G': 15,
        'H': 18,
        'I': 45,
        'J': 25,
        'K': 25,
        'L': 25,
        'M': 18,
        'N': 18,
        'O': 40,
        'P': 40,
    }

    for col, width in columnas.items():
        ws.column_dimensions[col].width = width

    # HOJA DE PROGRAMAS

    ws_programas = wb.create_sheet(
       title="PROGRAMAS"
    )

    ws_programas.append([
       "CLAVE",
        "NOMBRE DEL PROGRAMA"
    ])

    for cell in ws_programas[1]:

        cell.font = Font(
          bold=True,
          color="FFFFFF"
        )

        cell.fill = PatternFill(
           start_color="0d6efd",
           end_color="0d6efd",
           fill_type="solid"
        )

    programas = ProgramaBeca.objects.filter(
       activo=True
    ).order_by('clave')

    for programa in programas:

      ws_programas.append([
        programa.clave,
        programa.nombre
    ])

    ws_programas.column_dimensions['A'].width = 25
    ws_programas.column_dimensions['B'].width = 50

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    response['Content-Disposition'] = (
        'attachment; filename="Plantilla_Importacion_Expedientes.xlsx"'
    )

    wb.save(response)

    return response


 #CABBS
@login_required
def lista_cabbs(request):

    cabbs = CABB.objects.filter(
        activo=True
    ).prefetch_related(
        'localidades',
        'localidades__municipio',
        'localidades__escuelas'
    ).annotate(
        total_localidades=Count(
            'localidades',
            distinct=True
        ),
        total_municipios=Count(
            'localidades__municipio',
            distinct=True
        ),
        total_escuelas=Count(
            'localidades__escuelas',
            distinct=True
        )
    ).order_by('nombre')

    for cabb in cabbs:

        cabb.localidades_detalle = []

        for localidad in cabb.localidades.all().order_by(
            'municipio__nombre',
            'nombre'
        ):

            cabb.localidades_detalle.append({
                'municipio': localidad.municipio.nombre,
                'nombre': localidad.nombre,
                'total_escuelas': localidad.escuelas.count()
            })

    return render(request, 'cabbs/lista_cabbs.html', {
        'cabbs': cabbs
    })

@user_passes_test(es_bibliotecario)
def registrar_cabb(request):

    if request.method == 'POST':
        form = CABBForm(request.POST, request.FILES)

        if form.is_valid():
            form.save()
            messages.success(request, 'CABB registrado correctamente.')
            return redirect('lista_cabbs')

    else:
        form = CABBForm()

    return render(request, 'cabbs/form_cabb.html', {
        'form': form,
        'titulo': 'Registrar CABB'
    })

@user_passes_test(es_bibliotecario)
def editar_cabb(request, id):

    cabb = get_object_or_404(CABB, id=id)

    if request.method == 'POST':

        form = CABBForm(
            request.POST,
            request.FILES,
            instance=cabb
        )

        if form.is_valid():
            form.save()
            messages.success(
                request,
                'CABB actualizado correctamente.'
            )
            return redirect('lista_cabbs')

    else:

        form = CABBForm(instance=cabb)

    return render(request, 'cabbs/form_cabb.html', {
        'form': form,
        'titulo': 'Editar CABB'
    })






@user_passes_test(es_bibliotecario)
def importar_curps_excel(request):

    if request.method == 'POST':

        form = ImportarCurpForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            try:

                archivo = request.FILES['archivo_excel']

                wb = openpyxl.load_workbook(archivo)
                ws = wb.active

                actualizados = 0
                no_encontrados = 0
                errores = 0

                for row in ws.iter_rows(
                    min_row=2,
                    values_only=True
                ):

                    try:

                        numero_tarjeta = (
                            str(row[0]).strip()
                            if row[0] else ""
                        )

                        curp = (
                            str(row[1]).strip().upper()
                            if row[1] else ""
                        )

                        if not numero_tarjeta or not curp:
                            errores += 1
                            continue

                        libro = Libro.objects.filter(
                            numero_tarjeta_bancaria=numero_tarjeta
                        ).first()

                        if not libro:
                            no_encontrados += 1
                            continue

                        if libro.curp:
                            continue

                        libro.curp = curp
                        libro.save()

                        actualizados += 1

                    except Exception:
                        errores += 1

                messages.success(
                    request,
                    f'CURP actualizadas: {actualizados}, '
                    f'No encontrados: {no_encontrados}, '
                    f'Errores: {errores}'
                )

                return redirect('lista_libros')

            except Exception as e:

                messages.error(
                    request,
                    f'Error al importar archivo: {e}'
                )

    else:

        form = ImportarCurpForm()

    return render(
        request,
        'expedientes/importar_curps.html',
        {
            'form': form
        }
    )
@login_required
@user_passes_test(es_bibliotecario)
def bitacora_general(request):

    movimientos = BitacoraMovimiento.objects.select_related(
        'usuario',
        'libro'
    ).all()[:200]

    return render(
        request,
        'bitacora/general.html',
        {
            'movimientos': movimientos
        }
    )
